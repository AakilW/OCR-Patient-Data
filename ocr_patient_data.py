import os
import pdfplumber
import openpyxl

# Excel file path
excel_path = r"C:\Users\User\Documents\Patient_Data.xlsx"

# List of folders containing PDFs
pdf_folders = [
    r"C:\Users\User\Documents\Requisition\Folder1",
    r"C:\Users\User\Documents\Requisition\Folder2",
    r"C:\Users\User\Documents\Requisition\Folder3",
    r"C:\Users\User\Documents\Requisition\Folder4"
]

# Load Excel file
wb = openpyxl.load_workbook(excel_path)
ws = wb.active  # Assuming data is in the first sheet

# Read Accession# from Column B
accession_dict = {}  # Store Accession# to Row mapping
for row in range(2, ws.max_row + 1):  # Skipping header (row 1)
    accession = ws.cell(row=row, column=2).value
    if accession:
        accession_dict[str(accession)] = row  # Store as string for matching

# Function to extract required fields from PDF
def extract_patient_info(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                lines = text.split("\n")  # Split by new lines

                extracted_info = {
                    "FullName": lines[5] if len(lines) > 5 else "",
                    "DOB": lines[6] if len(lines) > 6 else "",
                    "Gender": lines[7] if len(lines) > 7 else "",
                    "RefferingProvider": lines[8] if len(lines) > 8 else "",
                    "Facility": lines[9] if len(lines) > 9 else "",
                    "InsuranceProvider": lines[10] if len(lines) > 10 else "",
                    "InsuranceID": lines[11] if len(lines) > 11 else "",
                    "BillingStatus": lines[12] if len(lines) > 12 else "",
                    "Biller": lines[13] if len(lines) > 13 else ""
                }
                return extracted_info
    return None  # Return None if no info found

# Process PDFs from all folders
for pdf_folder_path in pdf_folders:
    if not os.path.exists(pdf_folder_path):
        print(f"⚠️ Folder not found: {pdf_folder_path}")
        continue  # Skip if folder doesn't exist

    for filename in os.listdir(pdf_folder_path):
        if filename.endswith(".pdf"):
            accession_num = filename.replace(".pdf", "")  # Assuming filename = Accession#
            pdf_path = os.path.join(pdf_folder_path, filename)

            if accession_num in accession_dict:
                row = accession_dict[accession_num]
                patient_info = extract_patient_info(pdf_path)

                if patient_info:
                    ws.cell(row=row, column=4, value=patient_info["FullName"])
                    ws.cell(row=row, column=5, value=patient_info["DOB"])
                    ws.cell(row=row, column=6, value=patient_info["Gender"])
                    ws.cell(row=row, column=8, value=patient_info["RefferingProvider"])
                    ws.cell(row=row, column=9, value=patient_info["Facility"])
                    ws.cell(row=row, column=10, value=patient_info["InsuranceProvider"])
                    ws.cell(row=row, column=11, value=patient_info["InsuranceID"])
                    ws.cell(row=row, column=16, value=patient_info["BillingStatus"])
                    ws.cell(row=row, column=18, value=patient_info["Biller"])
                    print(f"✅ Updated: {accession_num} → {patient_info}")

# Save the updated Excel file
wb.save(excel_path)
print("\n✅ Excel file updated successfully with Patient Info.")
